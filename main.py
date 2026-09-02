import os, re, json, math, csv, threading
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.request import Request, urlopen

from kivy.app import App
from kivy.clock import Clock
from kivy.core.image import Image as CoreImage
from kivy.metrics import dp
from kivy.properties import StringProperty, NumericProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.progressbar import ProgressBar
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.graphics import Color, Rectangle

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_OK = True
except Exception:
    PIL_OK = False

RANKING_URL = "https://raw.githubusercontent.com/pvpoke/pvpoke/master/src/data/rankings/all/overall/rankings-{cp}.json"
POKEMON_URL = "https://raw.githubusercontent.com/pvpoke/pvpoke/master/src/data/gamemaster/pokemon.json"
TR_MOVES_URL = "https://raw.githubusercontent.com/WatWowMap/pogo-data-api/main/data/v1/translations/tr/moves.json"
POGO_MOVES_URL = "https://raw.githubusercontent.com/WatWowMap/pogo-data-api/main/data/v1/moves.json"
SPRITE_URL = "https://img.pokemondb.net/sprites/home/normal/{slug}.png"
POKEAPI_URL = "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/other/official-artwork/{id}.png"
TYPE_ICON_BASE = "https://raw.githubusercontent.com/WatWowMap/wwm-uicons/main/type/"
LEAGUES = {"Great League": 1500, "Ultra League": 2500, "Master League": 10000}
TYPE_TR = {"normal":"Normal","fire":"Ateş","water":"Su","electric":"Elektrik","grass":"Çim","ice":"Buz","fighting":"Dövüş","poison":"Zehir","ground":"Yer","flying":"Uçan","psychic":"Psişik","bug":"Böcek","rock":"Kaya","ghost":"Hayalet","dragon":"Ejderha","dark":"Karanlık","steel":"Çelik","fairy":"Peri"}
TYPE_ICON_ID = {"normal":1,"fighting":2,"flying":3,"poison":4,"ground":5,"rock":6,"bug":7,"ghost":8,"steel":9,"fire":10,"water":11,"grass":12,"electric":13,"psychic":14,"ice":15,"dragon":16,"dark":17,"fairy":18}
FALLBACK_TR = {"ROLLOUT":"Yuvarlanma","BODY_SLAM":"Vücut Çarpması","SHADOW_BALL":"Gölge Topu","EARTHQUAKE":"Deprem","HYPER_BEAM":"Hiper Işın","SOLAR_BEAM":"Güneş Işını"}
CPM_VALUES=[0.094,0.135137432,0.16639787,0.192650919,0.21573247,0.236572661,0.25572005,0.273530381,0.29024988,0.306057377,0.3210876,0.335445036,0.34921268,0.362457751,0.37523559,0.387592406,0.39956728,0.411193551,0.42250001,0.432926419,0.44310755,0.453059958,0.46279839,0.472336083,0.48168495,0.4908558,0.49985844,0.508701765,0.51739395,0.525942511,0.53435433,0.542635767,0.55079269,0.558830576,0.56675452,0.574569153,0.58227891,0.589887917,0.59740001,0.604818814,0.61215729,0.619399365,0.62656713,0.633644533,0.64065295,0.647576426,0.65443563,0.661214806,0.667934,0.674577537,0.68116492,0.687680648,0.69414365,0.700538673,0.70688421,0.713164996,0.71939909,0.725571552,0.7317,0.734741009,0.73776948,0.740785574,0.74378943,0.746781211,0.74976104,0.752729087,0.75568551,0.758630378,0.76156384,0.764486065,0.76739717,0.770297266,0.7731865,0.776064962,0.77893275,0.781790055,0.78463697,0.787473578,0.79030001,0.79280395,0.79530001,0.79780391,0.80030001,0.80280389,0.80530001,0.80780390,0.81030001,0.81280390,0.81530001,0.81780390,0.82030001,0.82280390,0.82530001,0.82780390,0.83030001,0.83280390,0.83530003,0.83780375,0.84030003,0.84280373,0.84530002]
CPM={1.0+i*.5:v for i,v in enumerate(CPM_VALUES)}; LEVELS=list(CPM.keys())

def get_json(url):
    req = Request(
        url,
        headers={"User-Agent": "SinKa-PvP-100-Android/1.0"}
    )
    ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    with urlopen(req, timeout=30, context=ssl_ctx) as r:
        return json.loads(r.read().decode("utf-8"))
def calc_cp(base, ivs, level):
    cpm=CPM[level]; a=(base['atk']+ivs[0])*cpm; d=(base['def']+ivs[1])*cpm; h=(base['hp']+ivs[2])*cpm
    return math.floor(a*math.sqrt(d)*math.sqrt(h)/10)

def stat_product(base,ivs,level):
    cpm=CPM[level]; a=(base['atk']+ivs[0])*cpm; d=(base['def']+ivs[1])*cpm; h=math.floor((base['hp']+ivs[2])*cpm); return a*d*h

def rank1_iv(base,limit):
    if limit>=10000: return (15,15,15),50.0,calc_cp(base,(15,15,15),50.0)
    best=None; bestv=(15,15,15); bestl=1.0; bestcp=0
    for ai in range(16):
      for di in range(16):
       for hi in range(16):
        lo,hh=0,len(LEVELS)-1; legal=-1
        while lo<=hh:
            mid=(lo+hh)//2; lv=LEVELS[mid]
            if calc_cp(base,(ai,di,hi),lv)<=limit: legal=mid; lo=mid+1
            else: hh=mid-1
        if legal<0: continue
        lv=LEVELS[legal]; cp=calc_cp(base,(ai,di,hi),lv); cpm=CPM[lv]
        key=(stat_product(base,(ai,di,hi),lv),cp,(base['def']+di)*cpm,math.floor((base['hp']+hi)*cpm),-ai)
        if best is None or key>best: best=key; bestv=(ai,di,hi); bestl=lv; bestcp=cp
    return bestv,bestl,bestcp

def shadow(p): return str(p.get('speciesId','')).endswith('_shadow') or '(Shadow)' in str(p.get('speciesName',''))
def slug(name): return re.sub(r'[^a-z0-9]+','-',name.lower()).strip('-')

def move_translator(pogo,tr):
    mp={str(m.get('proto')).upper():int(m.get('moveId')) for m in pogo if m.get('proto') is not None and m.get('moveId') is not None}
    def f(mid):
        x=mp.get(str(mid).upper())
        if x is not None:
            v=tr.get('move_'+str(x))
            if v and not str(v).startswith('<<'): return v
        return FALLBACK_TR.get(str(mid).upper(),str(mid).replace('_',' ').title())
    return f

class BGLabel(Label):
    bg = StringProperty('#ffffff')
    def __init__(self, **kw):
        super().__init__(**kw); self.bind(pos=self._upd,size=self._upd,bg=self._upd); self._upd()
    def _upd(self,*a):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._rgba(self.bg)); self._rect=Rectangle(pos=self.pos,size=self.size)
    def _rgba(self,h):
        h=h.lstrip('#'); return tuple(int(h[i:i+2],16)/255 for i in (0,2,4))+(1,)

class SinKaAndroid(App):
    def build(self):
        self.title='SinKa PvP 100'
        self.pokemon={}; self.ranking={}; self.trans=lambda x:x; self.rows=[]; self.owned={}; self.images={}; self.type_images={}; self.evolutions={}
        self.league='Great League'; self.count=36; self.show_shadow=False; self.no_xl=False; self.no_special=False; self.busy=False
        self.root_box=BoxLayout(orientation='vertical',padding=dp(8),spacing=dp(6))
        self.build_controls(); self.build_table()
        threading.Thread(target=self.load_initial,daemon=True).start()
        return self.root_box

    def build_controls(self):
        title=Label(text='[b]SinKa PvP 100[/b]',markup=True,font_size=dp(25),size_hint_y=None,height=dp(42),color=(.05,.12,.3,1)); self.root_box.add_widget(title)
        sub=Label(text='PvPoke Overall • Türkçe güçler • Rank-1 IV • XL • Özel Güç',font_size=dp(13),size_hint_y=None,height=dp(25)); self.root_box.add_widget(sub)
        top=BoxLayout(size_hint_y=None,height=dp(46),spacing=dp(5))
        self.league_spinner=Spinner(text=self.league,values=list(LEAGUES),size_hint_x=.25); self.league_spinner.bind(text=self.on_league); top.add_widget(self.league_spinner)
        self.count_input=TextInput(text='36',input_filter='int',multiline=False,size_hint_x=.09,hint_text='1-100'); top.add_widget(self.count_input)
        b=Button(text='Filtrele',size_hint_x=.11); b.bind(on_release=lambda *_:self.refresh()); top.add_widget(b)
        r=Button(text='↻ Yenile',size_hint_x=.11); r.bind(on_release=lambda *_:self.reload()); top.add_widget(r)
        x=Button(text='Excel Kaydet',size_hint_x=.13); x.bind(on_release=lambda *_:self.save_excel()); top.add_widget(x)
        p=Button(text='Resim Kaydet',size_hint_x=.13); p.bind(on_release=lambda *_:self.save_a4()); top.add_widget(p)
        self.root_box.add_widget(top)
        filt=BoxLayout(size_hint_y=None,height=dp(38),spacing=dp(8))
        self.cb_shadow=CheckBox(size_hint_x=None,width=dp(28)); self.cb_shadow.bind(active=self.set_shadow); filt.add_widget(self.cb_shadow); filt.add_widget(Label(text="Shadow'ları göster",halign='left',text_size=(None,None)))
        self.cb_xl=CheckBox(size_hint_x=None,width=dp(28)); self.cb_xl.bind(active=self.set_xl); filt.add_widget(self.cb_xl); filt.add_widget(Label(text='XL gerektirmeyenleri göster'))
        self.cb_special=CheckBox(size_hint_x=None,width=dp(28)); self.cb_special.bind(active=self.set_special); filt.add_widget(self.cb_special); filt.add_widget(Label(text='Özel güç (★) gerektirmeyenleri göster'))
        self.root_box.add_widget(filt)
        self.status=Label(text='Veriler yükleniyor...',size_hint_y=None,height=dp(30),halign='left',text_size=(None,None)); self.root_box.add_widget(self.status)
        self.progress=ProgressBar(max=100,value=0,size_hint_y=None,height=dp(8)); self.root_box.add_widget(self.progress)

    def build_table(self):
        headers=['Rank','Görsel','Pokémon','Tip','Önceki Evrimler','PvPoke','Rank 1 IV','Lvl','CP','XL','Özel Güç','Güç 1','Güç 2','Güç 3','Mevcut']
        widths=[70,90,150,190,270,80,105,65,70,65,90,180,180,180,80]
        self.hscroll=ScrollView(do_scroll_x=True,do_scroll_y=True,bar_width=dp(10))
        self.table=GridLayout(cols=15,size_hint=(None,None),spacing=1,padding=1)
        self.table.width=dp(sum(widths)); self.table.bind(minimum_height=self.table.setter('height'))
        for h,w in zip(headers,widths): self.table.add_widget(BGLabel(text='[b]'+h+'[/b]',markup=True,font_size=dp(12),color=(1,1,1,1),bg='#091f4e',size_hint=(None,None),size=(dp(w),dp(48)),halign='center',valign='middle',text_size=(dp(w)-6,dp(48))))
        self.hscroll.add_widget(self.table); self.root_box.add_widget(self.hscroll)

    def set_shadow(self,inst,val): self.show_shadow=val
    def set_xl(self,inst,val): self.no_xl=val
    def set_special(self,inst,val): self.no_special=val
    def set_status(self,s): Clock.schedule_once(lambda *_: setattr(self.status,'text',s))
    def on_league(self,sp,text):
        if text!=self.league:
            self.league=text
            if text in self.ranking: self.refresh()
            else: threading.Thread(target=self.load_league,args=(text,),daemon=True).start()
    def reload(self):
        threading.Thread(target=self.load_initial,daemon=True).start()
    def load_initial(self):
        try:
            self.set_status('Ana veriler indiriliyor...')
            plist,tr,pogo=get_json(POKEMON_URL),get_json(TR_MOVES_URL),get_json(POGO_MOVES_URL)
            self.pokemon={p.get('speciesId'):p for p in plist if p.get('speciesId')}; self.trans=move_translator(pogo,tr)
            self.load_league('Great League')
        except Exception as e: self.set_status('Hata: '+str(e))
    def load_league(self,league):
        try:
            self.set_status(league+' verisi indiriliyor...'); self.ranking[league]=get_json(RANKING_URL.format(cp=LEAGUES[league])); Clock.schedule_once(lambda *_: self.refresh())
        except Exception as e: self.set_status('Lig verisi indirilemedi: '+str(e))

    def get_count(self):
        try: return max(1,min(100,int(self.count_input.text or 36)))
        except: return 36
    def refresh(self,*_):
        if not self.pokemon or self.league not in self.ranking: return
        self.count=self.get_count(); candidates=[p for p in self.ranking[self.league] if self.show_shadow or not shadow(p)]
        self.status.text=f'{self.league} • hesaplanıyor...'; self.progress.value=0
        threading.Thread(target=self.calculate,args=(candidates,),daemon=True).start()

    def calculate(self,candidates):
        out=[]; total=len(candidates); done=0
        for p in candidates:
            done+=1
            sid=p.get('speciesId',''); name=p.get('speciesName',sid); poke=self.pokemon.get(sid) or self.pokemon.get(sid[:-7]) if sid.endswith('_shadow') else self.pokemon.get(sid)
            if not poke or 'baseStats' not in poke: continue
            base=poke['baseStats']; ivs,lv,cp=rank1_iv(base,LEAGUES[self.league]); moves=list(p.get('moveset',[]) or p.get('moves',[]) or [])[:3]
            elite=set(poke.get('eliteMoves',[])); legacy=set(poke.get('legacyMoves',[])); special=elite|legacy
            m=[]; has=False
            for mv in moves:
                t=self.trans(mv); 
                if mv in special: t+='*'; has=True
                m.append(t)
            while len(m)<3:m.append('-')
            xl=lv>40
            if self.no_xl and xl: pass
            elif self.no_special and has: pass
            else:
                typ=self.get_types(poke); evo=self.get_evos(sid,poke)
                out.append({'rank':len(out)+1,'name':name,'types':typ,'evo':evo,'score':f"{float(p.get('score',0)):.1f}",'iv':f'{ivs[0]}/{ivs[1]}/{ivs[2]}','level':f'{lv:g}','cp':cp,'xl':xl,'special':has,'moves':m,'owned':self.owned.get(name,False)})
            if done%5==0 or done==total:
                val=done/max(1,total)*100; Clock.schedule_once(lambda _,v=val,d=done,t=total:self.set_progress(v,d,t))
            if len(out)>=self.count: break
        self.rows=out; Clock.schedule_once(lambda *_:self.render_rows())
    def set_progress(self,v,d,t): self.progress.value=v; self.status.text=f'Veriler işleniyor... {d}/{t}'

    def get_types(self,p):
        vals=p.get('types') or p.get('type') or p.get('typing') or []
        if isinstance(vals,str): vals=[x.strip() for x in re.split('[/,]',vals)]
        out=[]
        for x in vals:
            if isinstance(x,dict): x=x.get('name') or x.get('type',{}).get('name')
            if x:
                k=str(x).lower(); out.append(k)
        return out
    def get_evos(self,sid,poke):
        parent=(poke.get('family') or {}).get('parent'); prev=[]; seen=set()
        while parent and parent not in seen:
            seen.add(parent); q=self.pokemon.get(parent) or self.pokemon.get(parent[:-7] if parent.endswith('_shadow') else parent)
            if not q: break
            prev.append(q.get('name') or q.get('speciesName') or parent); parent=(q.get('family') or {}).get('parent')
        prev.reverse(); self.evolutions[poke.get('name') or poke.get('speciesName') or sid]=prev; return prev

    def clear_data_rows(self):
        while len(self.table.children)>15: self.table.remove_widget(self.table.children[0])
    def render_rows(self):
        self.clear_data_rows()
        for i,r in enumerate(self.rows): self.add_row(r,i)
        self.status.text=f'✓ {len(self.rows)} Pokémon hazır'
        self.load_images_async()

    def cell(self,text,w,color='#12223a',bg='#ffffff',bold=False,halign='center'):
        lab=BGLabel(text=('[b]'+str(text)+'[/b]') if bold else str(text),markup=bold,font_size=dp(11),color=self.hex_rgba(color),bg=bg,size_hint=(None,None),size=(dp(w),dp(62)),halign=halign,valign='middle',text_size=(dp(w)-8,dp(62)))
        return lab
    def hex_rgba(self,h):
        h=h.lstrip('#'); return tuple(int(h[i:i+2],16)/255 for i in (0,2,4))+(1,)
    def add_row(self,r,i):
        widths=[70,90,150,190,270,80,105,65,70,65,90,180,180,180,80]; bg='#ffffff' if i%2==0 else '#f1f4f8'; fg='#da2330' if (r['xl'] or r['special']) else '#00914c'
        self.table.add_widget(self.cell(r['rank'],70,bg=bg)); self.table.add_widget(self.make_poke_image(r['name'],90,bg)); self.table.add_widget(self.cell(r['name'],150,fg,bg,True,'left')); self.table.add_widget(self.type_widget(r['types'],190,bg,fg)); self.table.add_widget(self.evo_widget(r['evo'],270,bg,fg)); self.table.add_widget(self.cell(r['score'],80,bg=bg)); self.table.add_widget(self.cell(r['iv'],105,bg=bg)); self.table.add_widget(self.cell(r['level'],65,bg=bg)); self.table.add_widget(self.cell(r['cp'],70,bg=bg)); self.table.add_widget(self.cell('✖' if r['xl'] else '✓',65,'#da2330' if r['xl'] else '#00914c',bg,True)); self.table.add_widget(self.cell('✖' if r['special'] else '✓',90,'#da2330' if r['special'] else '#00914c',bg,True));
        for mv in r['moves']: self.table.add_widget(self.cell(mv,180,fg,bg,False,'left'))
        cur=self.cell('☑' if r['owned'] else '☐',80,'#00914c' if r['owned'] else '#556270',bg,True); cur.bind(on_touch_down=lambda inst,t,rr=r: self.toggle_owned(rr,t)); self.table.add_widget(cur)
    def toggle_owned(self,r,t):
        if t.is_double_tap or (t.button=='left' and t.pos[0]>=0):
            r['owned']=not r['owned']; self.owned[r['name']]=r['owned']; self.render_rows(); return True
        return False
    def make_poke_image(self,name,w,bg):
        holder=BGLabel(text='',bg=bg,size_hint=(None,None),size=(dp(w),dp(62)))
        def load(_):
            data=self.images.get(name)
            if data:
                try:
                    im=Image(texture=CoreImage(BytesIO(data),ext='png').texture,size_hint=(None,None),size=(dp(52),dp(52)),pos_hint={'center_x':.5,'center_y':.5}); holder.add_widget(im)
                except: pass
        if name in self.images: Clock.schedule_once(load)
        return holder
    def type_widget(self,keys,w,bg,fg):
        box=BGLabel(text=' / '.join(TYPE_TR.get(k,k.title()) for k in keys),bg=bg,color=self.hex_rgba(fg),font_size=dp(10),size_hint=(None,None),size=(dp(w),dp(62)),halign='left',valign='middle',text_size=(dp(w)-8,dp(62)))
        return box
    def evo_widget(self,names,w,bg,fg):
        return self.cell('Evrim yok' if not names else ' → '.join(names),w,fg,bg,False,'left')

    def fetch_img(self,name):
        try:
            u=SPRITE_URL.format(slug=slug(name)); req=Request(u,headers={'User-Agent':'SinKa'}); return urlopen(req,timeout=15).read()
        except: return None
    def load_images_async(self):
        names=[]
        for r in self.rows:
            names.append(r['name']); names.extend(r['evo'])
        names=list(dict.fromkeys(names))
        def work():
            with ThreadPoolExecutor(max_workers=8) as pool:
                fs={pool.submit(self.fetch_img,n):n for n in names}
                for f in as_completed(fs):
                    n=fs[f]
                    try:
                        d=f.result()
                        if d:self.images[n]=d
                    except: pass
            Clock.schedule_once(lambda *_:self.render_rows())
        threading.Thread(target=work,daemon=True).start()

    def save_a4(self):
        if not self.rows: self.status.text='Önce listeyi yükle.'; return
        if not PIL_OK: self.status.text='Pillow bulunamadı.'; return
        threading.Thread(target=self._save_a4_worker,daemon=True).start()
    def font(self,size,bold=False):
        paths=['/system/fonts/Roboto-Bold.ttf' if bold else '/system/fonts/Roboto-Regular.ttf','/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf']
        for p in paths:
            try:return ImageFont.truetype(p,size)
            except: pass
        return ImageFont.load_default()
    def _save_a4_worker(self):
        try:
            W,H=4200,2976; margin=42; header_h=455; table_header_h=110; footer_h=105; img=PILImage.new('RGBA',(W,H),(248,250,253,255)); d=ImageDraw.Draw(img); NAVY=(9,31,78); BLUE=(29,101,218); CYAN=(30,194,180); GOLD=(245,178,42); RED=(218,35,48); GREEN=(0,145,76); TEXT=(18,34,58); GRID=(211,220,231)
            fb=self.font(118,True); fp=self.font(62); fl=self.font(122,True); fc=self.font(44,True); fh=self.font(38,True); fcell=self.font(42); fsmall=self.font(31); fsmallb=self.font(31,True); foot=self.font(30)
            hx0,hy0=margin,margin; hx1,hy1=W-margin,margin+header_h; d.rectangle((hx0,hy0,hx1,hy1),fill='white',outline=NAVY,width=5); d.polygon([(780,hy0),(1030,hy0),(1280,hy1),(1000,hy1)],fill=(37,112,222)); d.polygon([(680,hy0),(850,hy0),(1100,hy1),(930,hy1)],fill=(21,69,155)); d.polygon([(3220,hy0),(3510,hy0),(3280,hy1),(3020,hy1)],fill=(31,193,178)); d.polygon([(3490,hy0),(3740,hy0),(3510,hy1),(3260,hy1)],fill=(14,75,153))
            d.text((95,92),'SINKA',fill=NAVY,font=fb); d.text((105,218),'PvP 100',fill=BLUE,font=fp); d.line((105,290,560,290),fill=CYAN,width=8)
            for sx,sr in [(1710,26),(1800,40),(1890,26)]: d.regular_polygon((sx,82,sr),5,fill=GOLD)
            d.line((1550,140,2050,140),fill=GOLD,width=5); d.text((1800,226),self.league.upper(),fill=NAVY,font=fl,anchor='mm'); d.rounded_rectangle((1630,305,1970,375),35,fill=NAVY); d.text((1800,340),f"{LEAGUES[self.league]} CP MAX",fill='white',font=fc,anchor='mm')
            rx=W-105; d.text((rx,92),'SinKa PvP 100',fill=NAVY,font=fsmallb,anchor='ra'); d.text((rx,138),'PvPoke Overall',fill=TEXT,font=fsmall,anchor='ra'); d.text((rx,180),datetime.now().strftime('%d %B %Y'),fill=TEXT,font=fsmall,anchor='ra'); d.line((rx-330,215,rx,215),fill=CYAN,width=5); d.text((rx,252),f'{len(self.rows)} Pokémon',fill=NAVY,font=fsmallb,anchor='ra'); d.text((rx,292),f"XL Filtresi: {'UYGULANDI' if self.no_xl else 'UYGULANMADI'}",fill=GREEN if self.no_xl else TEXT,font=fsmallb,anchor='ra'); d.text((rx,332),f"Shadow: {'DAHİL EDİLDİ' if self.show_shadow else 'DAHİL EDİLMEDİ'}",fill=RED if self.show_shadow else TEXT,font=fsmallb,anchor='ra'); d.text((rx,372),f"Özel Güç Filtresi: {'UYGULANDI' if self.no_special else 'UYGULANMADI'}",fill=GREEN if self.no_special else TEXT,font=fsmallb,anchor='ra')
            d.rectangle((hx0,hy1-52,hx1,hy1),fill=NAVY); d.text((95,hy1-26),'PvPoke Overall Sıralaması  •  Rank 1 IV  •  XL  •  Özel Güç',fill='white',font=fsmall,anchor='lm'); d.text((W-100,hy1-26),'SINKA',fill='white',font=fsmallb,anchor='rm')
            cols=[('Rank',90),('Görsel',125),('Pokémon',310),('Tip',280),('Önceki Evrimler',610),('PvPoke',140),('Rank 1 IV',220),('Lvl',105),('CP',120),('XL',105),('Özel Güç',155),('Güç 1',330),('Güç 2',330),('Güç 3',330),('Mevcut',125)]; usable=W-2*margin; scale=usable/sum(x[1] for x in cols); cols=[(n,int(w*scale)) for n,w in cols]; xpos=[margin]
            for _,w in cols:xpos.append(xpos[-1]+w)
            top=hy1+24; avail=H-top-footer_h-margin; rowh=max(74,int((avail-table_header_h)/max(1,len(self.rows)))); d.rectangle((margin,top,W-margin,top+table_header_h),fill=NAVY)
            for i,(lab,_) in enumerate(cols): x0,x1=xpos[i],xpos[i+1]; d.line((x1,top,x1,top+table_header_h),fill=(80,105,145),width=2); d.text(((x0+x1)//2,top+table_header_h/2),lab,fill='white',font=fh,anchor='mm')
            y=top+table_header_h
            def fit(t,mx,f):
                t=str(t or '');
                while d.textbbox((0,0),t,font=f)[2]>mx and len(t)>1:t=t[:-1]
                return t if d.textbbox((0,0),t,font=f)[2]<=mx else t+'…'
            for idx,r in enumerate(self.rows):
                bg=(255,255,255) if idx%2==0 else (245,248,252); fg=RED if (r['xl'] or r['special']) else GREEN; d.rectangle((margin,y,W-margin,y+rowh),fill=bg,outline=GRID,width=1)
                vals=[r['rank'],'',r['name'],'', ' → '.join(r['evo']) if r['evo'] else 'Evrim yok',r['score'],r['iv'],r['level'],r['cp'],'','',r['moves'][0],r['moves'][1],r['moves'][2],'☑' if r['owned'] else '☐']
                for ci in range(15):
                    x0,x1=xpos[ci],xpos[ci+1]; d.line((x1,y,x1,y+rowh),fill=GRID,width=1)
                    if ci in (1,3,9,10):continue
                    tx=fit(vals[ci],x1-x0-22, fcell if ci!=2 else fsmallb); d.text((x0+12 if ci in (2,4,11,12,13) else (x0+x1)/2,y+rowh/2),tx,fill=fg if ci in (2,4,11,12,13) else TEXT,font=tx and (fsmallb if ci==2 else fcell),anchor='lm' if ci in (2,4,11,12,13) else 'mm')
                data=self.images.get(r['name'])
                if data:
                    try: pi=PILImage.open(BytesIO(data)).convert('RGBA'); pi.thumbnail((rowh-8,rowh-8),PILImage.LANCZOS); img.alpha_composite(pi,(int((xpos[1]+xpos[2]-pi.width)/2),int(y+(rowh-pi.height)/2)))
                    except:pass
                if not r['evo']:d.text((xpos[4]+10,y+rowh/2),'Evrim yok',fill=fg,font=fsmallb,anchor='lm')
                for ci,val in ((9,r['xl']),(10,r['special'])):
                    cx=(xpos[ci]+xpos[ci+1])/2; cy=y+rowh/2
                    if not val:d.line((cx-15,cy,cx-4,cy+13),fill=GREEN,width=7); d.line((cx-4,cy+13,cx+18,cy-15),fill=GREEN,width=7)
                    else:d.line((cx-13,cy-13,cx+13,cy+13),fill=RED,width=7); d.line((cx+13,cy-13,cx-13,cy+13),fill=RED,width=7)
                y+=rowh
            fy=H-margin-footer_h+20; d.line((margin+20,fy,W-margin-20,fy),fill=CYAN,width=5); d.text((W//2-480,fy+42),'www.sinka.com.tr',fill=NAVY,font=foot,anchor='mm'); d.text((W//2-20,fy+42),'|',fill=CYAN,font=foot,anchor='mm'); d.text((W//2+180,fy+42),'SinKa PvP 100',fill=NAVY,font=foot,anchor='mm'); d.text((W//2+430,fy+42),'|',fill=CYAN,font=foot,anchor='mm'); d.text((W//2+700,fy+42),datetime.now().strftime('%d %B %Y'),fill=NAVY,font=foot,anchor='mm')
            out=os.path.join(self.user_data_dir,f'SinKa_PvP100_{self.league.replace(" ","_")}.png'); img.convert('RGB').save(out,'PNG',dpi=(180,180)); Clock.schedule_once(lambda *_:setattr(self.status,'text','✓ A4 pafta kaydedildi: '+out))
        except Exception as e: Clock.schedule_once(lambda *_:setattr(self.status,'text','PNG hatası: '+str(e)))

    def save_excel(self):
        if not self.rows:self.status.text='Önce listeyi yükle.';return
        try:
            from openpyxl import Workbook
            wb=Workbook(); ws=wb.active; ws.title='SinKa PvP 100'; headers=['Rank','Görsel','Pokemon','Tip','Önceki Evrimler','PvPoke','Rank 1 IV','Lvl','CP','XL','Özel Güç','Güç 1','Güç 2','Güç 3','Mevcut']; ws.append(headers)
            for r in self.rows: ws.append([r['rank'],'',r['name'],' / '.join(TYPE_TR.get(k,k.title()) for k in r['types']),' → '.join(r['evo']) if r['evo'] else 'Evrim yok',r['score'],r['iv'],r['level'],r['cp'],'✖' if r['xl'] else '✓','✖' if r['special'] else '✓',*r['moves'],'☑' if r['owned'] else '☐'])
            out=os.path.join(self.user_data_dir,f'SinKa_PvP100_{self.league.replace(" ","_")}.xlsx'); wb.save(out); self.status.text='✓ Excel kaydedildi: '+out
        except Exception as e:self.status.text='Excel hatası: '+str(e)

if __name__=='__main__': SinKaAndroid().run()
